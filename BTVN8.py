import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "http://books.toscrape.com/catalogue/"

RATING_MAP = {
    "One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5
}

books = []

for page in range(1, 6):
    url = f"{BASE_URL}page-{page}.html"
    res = requests.get(url)
    res.encoding = "utf-8"
    soup = BeautifulSoup(res.text, "html.parser")

    for article in soup.select("article.product_pod"):
        ten_sach   = article.select_one("h3 a")["title"]
        gia        = article.select_one("p.price_color").text.strip().replace("Â", "").replace("£", "")
        rating_cls = article.select_one("p.star-rating")["class"][1]
        danh_gia   = RATING_MAP.get(rating_cls, 0)
        ton_kho    = article.select_one("p.availability").text.strip()

        books.append({
            "Ten Sach":   ten_sach,
            "Gia (GBP)":  float(gia),
            "Danh Gia":   danh_gia,
            "Tinh Trang": ton_kho
        })

df = pd.DataFrame(books)

output_path = "danh_sach_sach.xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Danh sach Sach", index=False)

wb = load_workbook(output_path)
ws = wb["Danh sach Sach"]

header_fill   = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
center_align  = Alignment(horizontal="center", vertical="center")
left_align    = Alignment(horizontal="left",   vertical="center")
thin          = Side(style="thin", color="BFBFBF")
border        = Border(left=thin, right=thin, top=thin, bottom=thin)

col_widths = [55, 12, 12, 18]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.row_dimensions[1].height = 22
for cell in ws[1]:
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = border

even_fill = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    fill = even_fill if row_idx % 2 == 0 else None
    for col_idx, cell in enumerate(row, start=1):
        cell.border = border
        cell.font   = Font(name="Arial", size=10)
        if fill:
            cell.fill = fill
        if col_idx == 1:
            cell.alignment = left_align
        else:
            cell.alignment = center_align

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(output_path)
print(f"Xong! Da luu {len(df)} sach vao '{output_path}'")